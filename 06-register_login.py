'''
1.代码自动读取测试数据====一般测试用例都放excel，需要导入操作excel的库
    自动化回写数据
2. 安装第三方库：pip install openpyxl
3. 导入第三方库：import openpyxl

excel表格常用操作：
1. 工作簿对象
2. 表单对象--sheet：工作簿对象.[表单名称]
3. 单元格对象--cell：表单对象.cell(row, column)
4. 获取值--cell.value()

========================
接口自动化测试流程：
1. 编写接口测试用例
2. Python代码读取测试用例里的数据----read_data()
3. requests库发送接口请求----api_request()
4. 执行结果 VS 预期结果 比对是否一致
5. 把比对结果回写到测试用例excel里面----write_result()
==========================

========================
函数实现流程：
1. 功能实现
2. 参数化
3. 判断是否需要返回值
==========================

'''

import openpyxl
import requests
import pprint

# keyword-列名
def get_column(file_name, sheet_name, keyword):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    max_column = sheet.max_column
    for col_index in range(1, max_column, 1):
        if keyword == sheet.cell(row=1, column=col_index).value:
            return col_index

# 1. 读取测试数据： file_name-用例名称， sheet_name-表格名称
def read_data(file_name, sheet_name):

    case_list = []
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    
    # index从2开始，因为用例设计得第一行为标题
    for i in range(2, max_row+1):

        # # 获取单元格值： 单元格对象.value()--column变量传入
        # column_caseid = get_column(file_name, sheet_name, "case_id")
        # column_url = get_column(file_name, sheet_name, "url")
        # column_data = get_column(file_name, sheet_name, "data")
        # column_exp = get_column(file_name, sheet_name, "expected")
        #
        # # 获取单元格值： 单元格对象.value()
        # case = dict(
        #     case_id = sheet.cell(row=i, column=column_caseid).value,
        #     url = sheet.cell(row=i, column=column_url).value,
        #     data = sheet.cell(row=i, column=column_data).value,
        #     expected = sheet.cell(row=i, column=column_exp).value)

        # 获取单元格值： 单元格对象.value()--column常量传入，必须与用例强相关
        case = dict(
            case_id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,
            data=sheet.cell(row=i, column=6).value,
            expected=sheet.cell(row=i, column=7).value)

        # 存储所有的用例数据到列表中
        case_list.append(case)

    return case_list


# 2. 发送接口请求： api_url-请求接口的url路径， api_data-接口数据
def api_request(api_url, api_data):
    api_header = {
        "X-Lemonban-Media-Type": "lemonban.v2",
        "Content-Type": "application/json"
    }

    response = requests.post(
        url=api_url,
        json=api_data,
        headers=api_header)

    return response.json()


# 3. 回写数据：file_name-用例名称， sheet_name-表格名称，row-行，column-列，final_result-最终结果
def write_result(file_name, sheet_name, row, column, final_result):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    sheet.cell(row=row, column=column).value = final_result

    # 调用save保存
    wb.save(file_name)


# 4. 结果断言：file_name-用例名称， sheet_name-表格名称
def execute_func(file_name, sheet_name):
    # 1-获取测试数据
    cases = read_data(file_name, sheet_name)
    for case in cases:
        case_id = case.get("case_id")
        url = case.get("url")
        data = case.get("data")
        expected = case.get("expected")

        # 从excel读取的数据是字符串的格式，必须转换成字典
        # 字符串（内含字典）转换成字典：用eval()函数
        # eval(): 去字符串引号功能，取出串里面的真实python表达式以及相对应的数据类型
        data = eval(data)
        expected = eval(expected)

        # 2-发送接口请求，并保存响应结果
        real_result = api_request(url, data)

        # 3-结果比对，添加断言
        real_msg = real_result.get("msg")
        expected_msg = expected.get("msg")

        # print("执行结果是：{}".format(real_msg))
        # print("预期结果是：{}".format(expected_msg))

        if real_msg == expected_msg:
            final_result = "Success"
            print("第{}条用例测试--通过".format(case_id))
        else:
            final_result = "Failed"
            print("第{}条用例测试--不通过".format(case_id))

        # 4-回写数据
        write_result(file_name, sheet_name, case_id+1, 8, final_result)

# 5. 函数调用
print("*" * 5, "register测试结果如下:", "*" * 5)
execute_func("test_case_api.xlsx", "register")

print("*" * 5, "login测试结果如下:", "*" * 5)
execute_func("test_case_api.xlsx", "login")
